import pandas as pd
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import os
from sigfig import round as sround


def chunks(lst, n):
    """Yield successive n-sized chunks from lst."""
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def plot_barh(series, title=None, sort=True, add_count=True, text_color='black', plot_dir=None, total_count=None,
              figsize=None):
    '''
    Take a series and plot a horizontal bar plot
    :param series: Series to plot.
    :param add_count: True to add the text count.
    :param text_color: Count text font color.
    :param plot_dir: String to indicate where to save the barh plot. If none, the plot is just shown.
    :return:
    '''
    fig, ax = plt.subplots(figsize=figsize)
    if sort:
        series = series.sort_values()
    series.plot.barh(ax=ax)
    max_value = series.max()
    if add_count:
        if total_count:
            for i, v in enumerate(series):
                ax.text(max_value*0.01, i-0.01, str(f'{sround(v, sigfigs=3)} / {sround(v/total_count*100, decimals=1)}%'),
                        color=text_color)
        else:
            for i, v in enumerate(series):
                ax.text(max_value*0.01, i-0.01, str(f'{sround(v, sigfigs=2)}'), color=text_color)
    plt.title(title)
    if plot_dir:
        plt.savefig(plot_dir, bbox_inches='tight')
    else:
        plt.show()
    plt.close()


def create_results_directory(results_directory, folders=None, excels=None):
    if os.path.exists(results_directory):
        expand = 1
        while True:
            expand += 1
            new_results_directory = results_directory + ' - ' + str(expand)
            if os.path.exists(new_results_directory):
                continue
            else:
                results_directory = new_results_directory
                break

    os.mkdir(results_directory)
    if folders:
        for item in folders:
            os.mkdir(results_directory + '/' + item)

    if excels:
        for item in excels:
            if item[-5:] != '.xlsx':
                item = item + '.xlsx'
            excel_name = results_directory + '/' + item
            wb = openpyxl.Workbook()
            wb.save(excel_name)
            wb.close()

    print('Creating new results directory: {}'.format(results_directory))
    return results_directory


def print_array_to_excel(array, first_cell, ws, axis=2):
    '''
    Print an np array to excel using openpyxl
    :param array: np array
    :param first_cell: first cell to start dumping values in
    :param ws: worksheet reference. From openpyxl, ws=wb[sheetname]
    :param axis: to determine if the array is a col vector (0), row vector (1), or 2d matrix (2)
    '''
    if isinstance(array, (list,)):
        array = np.array(array)
    shape = array.shape
    if axis == 0:
        # Treat array as col vector and print along the rows
        array=array.flatten()  # Flatten in case the input array is a nx1 ndarry which acts weird
        for i in range(shape[0]):
            j = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i]
    elif axis == 1:
        # Treat array as row vector and print along the columns
        array=array.flatten()  # Flatten in case the input array is a 1xn ndarry which acts weird
        for j in range(shape[0]):
            i = 0
            ws.cell(i + first_cell[0], j + first_cell[1]).value = array[j]
    elif axis == 2:
        # If axis==2, means it is a 2d array
        for i in range(shape[0]):
            for j in range(shape[1]):
                ws.cell(i + first_cell[0], j + first_cell[1]).value = array[i, j]


def print_df_to_excel(df, ws, start_row=1, start_col=1, index=True, header=True):
    rows = list(dataframe_to_rows(df, index=index, header=header))
    rows.pop(1)
    for r_idx, row in enumerate(rows, start_row):
        skip_count = 0
        for c_idx, value in enumerate(row, start_col):
            if isinstance(value, str):
                if 'Unnamed' not in value:
                    ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
            else:
                ws.cell(row=r_idx - skip_count, column=c_idx, value=value)
        else:
            skip_count += 1


def create_excel_file(excel_name):
    while os.path.isfile(excel_name):
        expand = 1
        while True:
            expand += 1
            new_file_name = excel_name.split('.xlsx')[0] + ' - ' + str(expand) + '.xlsx'
            if os.path.isfile(new_file_name):
                continue
            else:
                excel_name = new_file_name
                break
    print('Writing into' + excel_name + '\n')
    wb = openpyxl.Workbook()
    wb.save(excel_name)
    return excel_name

